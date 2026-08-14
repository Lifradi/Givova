from datetime import datetime
import io
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import streamlit as st

# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(
    page_title="Comparador e Destacador de CT-es",
    page_icon="📊",
    layout="wide",
)

# --- VALIDAÇÃO DE SEGURANÇA DA SESSÃO ---
authentication_status = st.session_state.get("authentication_status")
name = st.session_state.get("name")

if not authentication_status:
    st.error(
        "🔒 Acesso Negado! Por favor, faça o login na página principal do Hub"
        " primeiro."
    )
    if st.button("Ir para o Login"):
        st.switch_page("app.py")
    st.stop()

st.sidebar.markdown(f"👤 **Logado como:** {name}")

st.title("📊 Comparador e Destacador de CT-es - PRO")
st.write(
    "Envie a **Planilha 1** (com múltiplas abas de dias de operação) e a **Planilha 2** "
    "(a planilha grande de referência com 1 aba). O sistema cruzará os CTes e "
    "destacará em verde as linhas correspondentes na Planilha 2."
)

col1, col2 = st.columns(2)
with col1:
    uploaded_p1 = st.file_uploader(
        "Envie a Planilha 1 (Múltiplas abas) (.xlsx)",
        type=["xlsx"],
        key="p1",
    )
with col2:
    uploaded_p2 = st.file_uploader(
        "Envie a Planilha 2 de Referência (Grande, 1 aba) (.xls ou .xlsx)",
        type=["xls", "xlsx"],
        key="p2",
    )


def limpar_cte(valor):
    """Padroniza o valor do CTe removendo sufixos e formatações indesejadas."""
    if pd.isna(valor):
        return ""
    val_str = str(valor).strip()
    val_str = re.sub(r"[-_\s]*cte\b", "", val_str, flags=re.IGNORECASE)
    if val_str.endswith(".0"):
        val_str = val_str[:-2]
    return val_str.strip()


if uploaded_p1 and uploaded_p2:
    if st.button("🚀 Processar e Destacar Planilha 2", type="primary"):
        with st.spinner("Lendo abas, cruzando dados e pintando a planilha..."):
            try:
                bytes_p1 = uploaded_p1.read()
                bytes_p2 = uploaded_p2.read()

                # 1. Varre TODAS as abas da Planilha 1 para extrair todos os CTes operacionais
                xls1 = pd.ExcelFile(io.BytesIO(bytes_p1))
                ctes_operacao = set()

                for sheet in xls1.sheet_names:
                    df_sheet = pd.read_excel(
                        io.BytesIO(bytes_p1), sheet_name=sheet, header=None
                    )
                    for col_idx in df_sheet.columns:
                        for val in df_sheet[col_idx]:
                            limpo = limpar_cte(val)
                            if limpo.isdigit() and len(limpo) >= 3:
                                ctes_operacao.add(limpo)

                # 2. Lê a Planilha 2 usando pandas (suporta tanto .xls quanto .xlsx perfeitamente)
                # Determinamos o engine automaticamente baseado na extensão ou conteúdo
                file_name_p2 = uploaded_p2.name.lower()
                engine_p2 = "xlrd" if file_name_p2.endswith(".xls") else None
                df_p2 = pd.read_excel(
                    io.BytesIO(bytes_p2), sheet_name=0, engine=engine_p2
                )

                # Identifica colunas de CTe/Conhecimento na Planilha 2
                colunas_alvo_indices = []
                for idx, col_nome in enumerate(df_p2.columns):
                    if any(
                        termo in str(col_nome).upper()
                        for termo in ["CONHEC", "CTE"]
                    ):
                        colunas_alvo_indices.append(idx)  # 0-based index para pandas

                # 3. Constrói um novo Workbook moderno em memória usando openpyxl para aplicar as formatações
                wb_out = Workbook()
                ws_out = wb_out.active
                ws_out.title = "Planilha 2 Destacada"

                # Escreve o cabeçalho
                ws_out.append(list(df_p2.columns))

                fill_verde = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )
                total_destaques = 0

                # 4. Percorre as linhas do DataFrame da Planilha 2, verifica os CTes e escreve formatado
                for row_idx, row in df_p2.iterrows():
                    linha_encontrada = False

                    # Verifica se o CTe está presente na linha nas colunas alvo (ou em todas se não houver alvo)
                    cols_a_checar = (
                        colunas_alvo_indices
                        if colunas_alvo_indices
                        else range(len(df_p2.columns))
                    )

                    for col_i in cols_a_checar:
                        val_celula = row.iloc[col_i]
                        if pd.notna(val_celula):
                            val_limpo = limpar_cte(val_celula)
                            if val_limpo in ctes_operacao:
                                linha_encontrada = True
                                break

                    # Prepara os valores para inserção na linha
                    valores_linha = [
                        ("" if pd.isna(v) else v) for v in row.values
                    ]
                    ws_out.append(valores_linha)
                    current_row_idx = ws_out.max_row

                    if linha_encontrada:
                        total_destaques += 1
                        for col_i in range(1, len(row.values) + 1):
                            ws_out.cell(
                                row=current_row_idx, column=col_i
                            ).fill = fill_verde

                # Salva o arquivo resultante em memória
                output = io.BytesIO()
                wb_out.save(output)
                processed_data = output.getvalue()

                st.success(
                    f"✨ Processo concluído! Foram destacadas {total_destaques}"
                    " linhas na Planilha 2."
                )

                st.download_button(
                    label="📥 Baixar Planilha 2 Destacada (.XLSX)",
                    data=processed_data,
                    file_name=(
                        "planilha_2_destacada_"
                        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    ),
                    mime=(
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    ),
                    use_container_width=True,
                )

            except Exception as e:
                st.error(f"❌ Erro crítico ao processar as planilhas: {e}")
else:
    st.info(
        "💡 Envie ambas as planilhas acima para habilitar o processamento."
    )
